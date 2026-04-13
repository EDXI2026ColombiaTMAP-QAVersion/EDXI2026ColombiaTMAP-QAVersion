"""
Extract April schedule from 2026 - Timing_Map_DXI (1).xlsx
and produce a JSON file ready to import into the Timing Map app.
"""
import openpyxl
import json
import re
from datetime import date

EXCEL_FILE = "2026 - Timing_Map_DXI (1).xlsx"
SHEET_NAME = "Timing APR"
OUTPUT_FILE = "april_import.json"
YEAR = 2026

# Day columns layout in the Excel (0-indexed):
# Each week has 5 days, all starting at the same cols relative to the week-header row.
DAY_COLS = [3, 35, 67, 99, 131]  # 0-indexed column for each day's start

# App slots: 7:00 to 16:30, 30-min increments = 20 slots (index 0..19)
# Slot s → hour = 7 + s//2, minute = (s%2)*30
# Excel col offset from day_start:
#   hour h, minute 0 → offset = (h-6)*2
#   hour h, minute 30 → offset = (h-6)*2 + 1
# Inverse: given offset o → hour = 6 + o//2, half = o%2
#   → app slot = (hour-7)*2 + half  (valid only if hour >= 7 and hour <= 16)

def excel_col_to_app_slot(day_start_col, col):
    offset = col - day_start_col
    if offset < 0:
        return None
    hour = 6 + offset // 2
    half = offset % 2
    if hour < 7 or hour > 16:
        return None
    if hour == 16 and half == 1:
        return None  # 16:30 is the last slot (index 19), allow it
    return (hour - 7) * 2 + half

def parse_day_number(day_header):
    """Extract day number from string like 'MONDAY 30', 'WEDNESDAY 1', 'MONDAY  13'"""
    m = re.search(r'\d+', str(day_header))
    return int(m.group()) if m else None

def resolve_date(week_index, day_pos, day_num):
    """
    Determine full date from week index (0=first), day position (0=Mon..4=Fri), day number.
    April 2026 starts on Wednesday → Mon/Tue of first week are March 30/31.
    """
    if week_index == 0 and day_num > 20:
        # March dates (30, 31)
        return date(YEAR, 3, day_num).isoformat()
    else:
        # Check if it could be May (last week, day < 5)
        # Last week of April 2026: Mon 27, Tue 28, Wed 29, Thu 30, Fri 1 (May)
        if week_index >= 4 and day_num < 5:
            return date(YEAR, 5, day_num).isoformat()
        return date(YEAR, 4, day_num).isoformat()

def normalize_brand(name):
    """Normalize brand name from Excel for use as app brand name."""
    if not name or not isinstance(name, str):
        return None
    name = name.strip()
    if not name or name in ('DAY', 'j', '|', '}', 'c', 'h'):
        return None
    if len(name) <= 2:
        return None
    # Skip color codes and non-brand noise
    if name.startswith('#') and len(name) == 7:
        return None
    return name

wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
ws = wb[SHEET_NAME]
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

# ── 1. Find week-header rows ───────────────────────────────────────────────────
week_starts = []  # list of (row_index_0based, week_index)
WEEK_LABELS = ['FIRST WEEK', 'SECOND WEEK', 'THIRD WEEK', 'FOURTH WEEK', 'FIFTH WEEK']
for i, row in enumerate(rows):
    for w, label in enumerate(WEEK_LABELS):
        if any(c == label for c in row if c):
            week_starts.append((i, w))
            break

print(f"Found {len(week_starts)} week blocks at rows: {[r+1 for r,_ in week_starts]}")

# ── 2. Parse each week block ───────────────────────────────────────────────────
all_brands = set()           # unique brand names seen
# assignments: { date_str: { member_name: [slot0, slot1, ..., slot19] } }
assignments = {}
members_seen = []            # ordered list

DAY_NAMES = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY']

for block_idx, (week_row_0, week_index) in enumerate(week_starts):
    # Row right after the week header has day names
    day_header_row = rows[week_row_0]

    # Collect day headers: find the text at or near each DAY_COL
    # day_header_row has "MONDAY XX" at one of the DAY_COLS
    day_dates = {}  # day_position (0-4) → date_str
    for day_pos, day_col in enumerate(DAY_COLS):
        # The day header might be at day_col or nearby; scan ±2 cells
        header_val = None
        for offset in [0, -1, 1, -2, 2]:
            c = day_col + offset
            if 0 <= c < len(day_header_row) and day_header_row[c]:
                v = str(day_header_row[c]).strip()
                if any(dn in v.upper() for dn in DAY_NAMES):
                    header_val = v
                    break
        if not header_val:
            continue
        day_num = parse_day_number(header_val)
        if day_num is None:
            continue
        date_str = resolve_date(week_index, day_pos, day_num)
        day_dates[day_pos] = date_str

    if not day_dates:
        print(f"  Week {week_index+1}: no day dates found, skipping")
        continue

    # Determine the row range: from week_row_0 + 2 (skip header+hour rows)
    # until the next week block starts
    if block_idx + 1 < len(week_starts):
        end_row = week_starts[block_idx + 1][0]
    else:
        end_row = len(rows)

    # Scan member rows in this block (rows from week_row_0+2 to end_row)
    # A member row has col 1 (0-indexed) with a name string
    for row_idx in range(week_row_0 + 2, end_row):
        row = rows[row_idx]
        if len(row) < 2:
            continue
        member_name = row[1]  # column B (0-indexed = 1)
        if not member_name or not isinstance(member_name, str):
            continue
        member_name = member_name.strip()
        if not member_name or member_name in ('DAY',):
            continue
        # Skip hour-header rows (col 1 is a number)
        if isinstance(row[1], (int, float)):
            continue
        # Skip rows that look like week labels
        if any(label in member_name.upper() for label in ['WEEK', 'TIMING']):
            continue

        if member_name not in [m for m in members_seen]:
            members_seen.append(member_name)

        # For each day, read the slot assignments
        for day_pos, date_str in day_dates.items():
            day_col = DAY_COLS[day_pos]
            if date_str not in assignments:
                assignments[date_str] = {}
            if member_name not in assignments[date_str]:
                assignments[date_str][member_name] = [None] * 20

            # Read each cell in this row within this day's column range
            # Day range: day_col to day_col+29
            day_end_col = day_col + 30
            for col_idx in range(day_col, min(day_end_col, len(row))):
                brand_raw = row[col_idx]
                brand = normalize_brand(brand_raw)
                if not brand:
                    continue
                slot = excel_col_to_app_slot(day_col, col_idx)
                if slot is None:
                    continue
                all_brands.add(brand)
                assignments[date_str][member_name][slot] = brand

print(f"\nMembers ({len(members_seen)}): {members_seen}")
print(f"\nBrands ({len(all_brands)}): {sorted(all_brands)}")
print(f"\nDates covered: {sorted(assignments.keys())}")

# ── 3. Output JSON ─────────────────────────────────────────────────────────────
output = {
    "members": members_seen,
    "brands": sorted(all_brands),
    "assignments": assignments
}

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n✓ Written to {OUTPUT_FILE}")
